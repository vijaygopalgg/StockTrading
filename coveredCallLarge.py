"""
stock_options_puller.py
-----------------------
Fetches S&P 500 + Dow Jones tickers, then for each stock pulls:
  - Current price (as of today)
  - ITM Covered Call: 1 strike ABOVE current price
      * Next week Friday expiry
      * Next-to-next week Friday expiry

Ticker source:
  Custom master list defined in MASTER_TICKERS (editable at top of file)

Usage:
  python stock_options_puller.py
  python stock_options_puller.py --output "C:\\Users\\vijay\\options_output.xlsx"
  python stock_options_puller.py --workers 8
"""

import argparse
import random
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Master ticker list ────────────────────────────────────────────────────────

MASTER_TICKERS = [
    "A","AAL","AAP","AAPL","ABBV","ABC","ABMD","ABNB","ABR","ABT","ACGL","ACN","ACV","ADBE",
    "ADI","ADM","ADP","ADSK","AEE","AEP","AES","AFL","AIG","AIZ","AJG","AKAM","ALB","ALGN",
    "ALK","ALL","ALLE","ALSN","AMAT","AMCR","AMD","AME","AMGN","AMP","AMT","AMZN","ANET",
    "ANSS","AON","AOS","APA","APD","APH","APLD","APO","APP","APTV","ARE","ARGX","ARM","ASML",
    "ATO","ATVI","AVB","AVGO","AVY","AWK","AXON","AXP","AZN","AZO","BA","BAC","BAH","BAX",
    "BBWI","BBY","BDX","BEN","BF-B","BIIB","BIO","BK","BKNG","BKR","BLK","BMY","BR","BRK-B",
    "BRO","BSX","BWA","BXP","C","CAG","CAH","CAKE","CARR","CAT","CAVA","CB","CBOE","CBRE",
    "CCEP","CCI","CCL","CDNS","CDW","CE","CEG","CELH","CF","CFG","CHD","CHRW","CHTR","CI",
    "CINF","CL","CLS","CLX","CMA","CMCSA","CME","CMG","CMI","CMS","CNC","CNP","COF","COIN",
    "COO","COP","COST","CPB","CPRT","CPT","CRL","CRM","CRS","CRWD","CSCO","CSGP","CSX","CTAS",
    "CTLT","CTRA","CTSH","CTVA","CUK","CVNA","CVS","CVX","CZR","D","DAL","DASH","DD","DDOG",
    "DE","DECK","DFS","DG","DGX","DHI","DHR","DIS","DISH","DKNG","DLR","DLTR","DOV","DOW",
    "DPZ","DRE","DRI","DTE","DUK","DVA","DVN","DXC","DXCM","EA","EBAY","ECL","ED","EFX","EIX",
    "EL","ELF","EME","EMN","EMR","ENPH","ENX","EOG","EPAM","EPD","EQIX","EQR","EQT","ES","ESS",
    "ETN","ETR","EVRG","EW","EXC","EXEL","EXPD","EXPE","EXR","F","FANG","FAST","FCNCA","FCX",
    "FDS","FDX","FE","FFIV","FI","FIG","FIS","FISV","FITB","FLT","FMC","FOX","FOXA","FRC",
    "FRO","FRSH","FRT","FTAI","FTNT","FTV","FUTU","GD","GE","GEHC","GEN","GFS","GILD","GIS",
    "GL","GLW","GM","GNRC","GOOG","GOOGL","GPC","GPN","GRAL","GRMN","GS","GWW","HAL","HAS",
    "HBAN","HCA","HD","HES","HIG","HII","HLT","HOLX","HON","HOOD","HPE","HPQ","HRL","HSIC",
    "HST","HSY","HTGC","HUM","HWM","IBKR","IBM","ICE","IDXX","IEX","IFF","ILMN","INCY","INSW",
    "INTC","INTU","INVH","IP","IPG","IQV","IR","IRM","ISRG","IT","ITW","IVZ","J","JBHT","JCI",
    "JEPI","JEPQ","JKHY","JNJ","JNPR","JPM","K","KDP","KEY","KEYS","KGC","KHC","KIM","KLAC",
    "KMB","KMI","KMX","KNSL","KO","KR","L","LBRT","LDOS","LEN","LH","LHX","LI","LIN","LKQ",
    "LLY","LMT","LNC","LNT","LOW","LRCX","LULU","LUV","LVS","LW","LYB","LYV","MA","MAA","MAR",
    "MARA","MAS","MCD","MCHP","MCK","MCO","MDLZ","MDT","MELI","MET","META","MGM","MHK","MKC",
    "MKTX","MLM","MMC","MMM","MNST","MO","MOD","MOH","MOS","MPC","MPWR","MRK","MRNA","MRO",
    "MRVL","MS","MSCI","MSFT","MSI","MSTR","MTB","MTD","MU","NCLH","NDAQ","NE","NEE","NEM",
    "NFLX","NGD","NI","NKE","NOC","NOV","NOW","NRG","NSC","NTAP","NTRS","NUE","NVDA","NVO",
    "NVR","NWL","NWS","NWSA","NXPI","O","ODFL","OGN","OKE","OMC","ON","ONON","ORCL","ORLY",
    "OXY","PANW","PAYC","PAYX","PCAR","PCG","PDD","PEAK","PEG","PEP","PERI","PFE","PFG","PG",
    "PGR","PH","PHM","PKG","PKI","PLD","PLTR","PM","PNC","PNR","PNW","POOL","PPG","PPL","PRU",
    "PSA","PSX","PTC","PWR","PXD","PYPL","QCOM","QLYS","QRVO","RCL","RDDT","RE","REG","REGN",
    "RELY","RF","RHI","RIOT","RJF","RL","RMBS","RMD","ROK","ROL","ROP","ROST","RSG","RTX",
    "SBAC","SBUX","SCHW","SEE","SHOP","SHW","SJM","SLB","SNA","SNPS","SO","SOFI","SPG","SPGI",
    "SPOT","SRE","STE","STT","STX","STZ","SWK","SWKS","SYF","SYK","SYY","T","TAP","TDG","TDW",
    "TDY","TEAM","TECH","TEL","TER","TEVA","TEX","TFC","TFX","TGT","TJX","TLRY","TMO","TMUS",
    "TPR","TRMB","TROW","TRV","TS","TSCO","TSLA","TSN","TT","TTD","TTWO","TXN","TXT","TYL",
    "UAL","UDR","UHS","ULTA","UNH","UNP","UPS","URI","USB","V","VFC","VLO","VMC","VNO","VRSK",
    "VRSN","VRT","VRTX","VTR","VTRS","VZ","WAB","WAT","WBA","WBD","WDAY","WDC","WEC","WELL",
    "WFC","WHD","WHR","WM","WMB","WMT","WRB","WRK","WST","WTW","WY","WYNN","XEL","XOM","XRAY",
    "XYL","YUM","ZBH","ZBRA","ZION","ZS","ZTS","ZWS",
]


# ── Date helpers ──────────────────────────────────────────────────────────────

def next_friday(from_date=None):
    d = from_date or date.today()
    days_ahead = (4 - d.weekday()) % 7
    if days_ahead == 0:
        days_ahead = 7
    return d + timedelta(days=days_ahead)


def get_expiry_dates():
    f1 = next_friday()
    f2 = next_friday(f1 + timedelta(days=1))
    return str(f1), str(f2)


# ── Ticker map builder ────────────────────────────────────────────────────────

def build_ticker_index_map():
    """Return {ticker: 'Custom'} for every ticker in MASTER_TICKERS."""
    print(f"  Using master ticker list ({len(MASTER_TICKERS)} tickers)")
    return {t: "Custom" for t in MASTER_TICKERS}


# ── Options fetcher ───────────────────────────────────────────────────────────

def get_itm_call(ticker_obj, expiry, current_price):
    """Return (strike, mid_price) for 1 strike above current price."""
    try:
        chain = ticker_obj.option_chain(expiry)
        calls = chain.calls.copy()
        calls = calls[calls["strike"] > current_price].sort_values("strike")
        if calls.empty:
            return None, None
        row    = calls.iloc[0]
        strike = float(row["strike"])
        bid    = float(row["bid"])       if row["bid"]       > 0 else None
        ask    = float(row["ask"])       if row["ask"]       > 0 else None
        last   = float(row["lastPrice"]) if row["lastPrice"] > 0 else None
        if bid and ask:
            mid = round((bid + ask) / 2, 2)
        elif last:
            mid = round(last, 2)
        else:
            mid = None
        return strike, mid
    except Exception:
        return None, None


# ── Rate-limit safe fetch helpers ────────────────────────────────────────────

def _is_rate_limit_error(e):
    msg = str(e).lower()
    return "too many requests" in msg or "rate limit" in msg or "429" in msg

def _fetch_with_retry(fn, retries=4, base_delay=5):
    """Call fn(), retrying on rate-limit errors with exponential backoff + jitter."""
    for attempt in range(retries):
        try:
            return fn()
        except Exception as e:
            if _is_rate_limit_error(e) and attempt < retries - 1:
                wait = base_delay * (2 ** attempt) + random.uniform(0, 2)
                time.sleep(wait)
            else:
                raise
    return None


# ── Per-ticker processing ─────────────────────────────────────────────────────

def process_ticker(ticker, index_label, expiry1, expiry2):
    # Small random jitter before every request to spread out parallel calls
    time.sleep(random.uniform(0.1, 0.6))

    row = {
        "Ticker": ticker,
        "Company Name": None,
        "Industry": None,
        "Earnings Date": None,
        "Earnings Soon": None,
        "Index": index_label,
        "Current Price": None,
        "Date": str(date.today()),
        "Next Friday Expiry": expiry1,
        "Next Friday Strike": None,
        "Next Friday Call Price ($)": None,
        "Next-Next Friday Expiry": expiry2,
        "Next-Next Strike": None,
        "Next-Next Call Price ($)": None,
        "NF Strike - Current Price": None,
        "N2F Strike - Current Price": None,
        "NF Signal": None,
        "N2F Signal": None,
        "Notes": "",
    }
    try:
        t = yf.Ticker(ticker)

        # ── Company Name, Industry & Earnings — consolidated into t.info ──
        # t.info is one HTTP call; t.calendar is another. We pull both but
        # wrap each separately so one failure doesn't block the other.
        try:
            info = _fetch_with_retry(lambda: t.info)
            row["Company Name"] = info.get("longName") or info.get("shortName") or None
            row["Industry"]     = info.get("industry") or info.get("sector") or None
        except Exception:
            pass  # non-fatal

        try:
            cal = _fetch_with_retry(lambda: t.calendar)
            ed  = cal.get("Earnings Date") if cal else None
            if ed:
                ed_date = ed[0].date() if hasattr(ed[0], "date") else ed[0]
                row["Earnings Date"] = str(ed_date)
                days_away = (ed_date - date.today()).days
                row["Earnings Soon"] = "⚠️ Yes" if 0 <= days_away <= 14 else "No"
        except Exception:
            pass  # non-fatal

        # ── Price ──
        price = round(float(_fetch_with_retry(lambda: t.fast_info.last_price)), 2)
        row["Current Price"] = price

        # ── Options ──
        available = _fetch_with_retry(lambda: t.options)
        if not available:
            row["Notes"] = "No options data"
            return row

        match1 = expiry1 if expiry1 in available else next(
            (e for e in available if e >= expiry1), None)
        if match1:
            row["Next Friday Expiry"] = match1
            row["Next Friday Strike"], row["Next Friday Call Price ($)"] = \
                get_itm_call(t, match1, price)

        match2 = expiry2 if expiry2 in available else next(
            (e for e in available if e > (match1 or expiry1)), None)
        if match2:
            row["Next-Next Friday Expiry"] = match2
            row["Next-Next Strike"], row["Next-Next Call Price ($)"] = \
                get_itm_call(t, match2, price)

        # Derived columns
        nf_strike  = row["Next Friday Strike"]
        n2f_strike = row["Next-Next Strike"]
        if nf_strike is not None:
            diff1 = round(nf_strike - price, 2)
            row["NF Strike - Current Price"] = diff1
            row["NF Signal"] = "Good" if diff1 > 0 else "Bad"
        if n2f_strike is not None:
            diff2 = round(n2f_strike - price, 2)
            row["N2F Strike - Current Price"] = diff2
            row["N2F Signal"] = "Good" if diff2 > 0 else "Bad"

    except Exception as e:
        row["Notes"] = f"Error: {e}"

    return row


# ── Excel output ──────────────────────────────────────────────────────────────

COLUMNS = [
    "Ticker", "Company Name", "Industry", "Earnings Date", "Earnings Soon",
    "Index", "Current Price", "Date",
    "Next Friday Expiry", "Next Friday Strike", "Next Friday Call Price ($)",
    "Next-Next Friday Expiry", "Next-Next Strike", "Next-Next Call Price ($)",
    "NF Strike - Current Price", "N2F Strike - Current Price",
    "NF Signal", "N2F Signal",
    "Notes",
]
COL_WIDTHS = [10, 30, 25, 14, 13, 8, 14, 12, 18, 18, 22, 22, 16, 22, 22, 22, 12, 12, 35]


def save_excel(rows, output_path):
    df = pd.DataFrame(rows)[COLUMNS]
    df.to_excel(str(output_path), index=False)

    wb = load_workbook(str(output_path))
    ws = wb.active

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    alt_fill    = PatternFill("solid", start_color="D6E4F0")
    price_fill  = PatternFill("solid", start_color="E2EFDA")
    both_fill   = PatternFill("solid", start_color="FFF2CC")
    err_fill    = PatternFill("solid", start_color="FCE4D6")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")
    left_align  = Alignment(horizontal="left", vertical="center")

    for i in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]
    ws.row_dimensions[1].height = 28

    good_fill = PatternFill("solid", start_color="C6EFCE")   # green  - Good
    bad_fill  = PatternFill("solid", start_color="FFC7CE")   # red    - Bad
    diff_fill = PatternFill("solid", start_color="EDEDED")   # grey   - diff columns

    # Column positions (1-based):
    # 1=Ticker, 2=Company Name, 3=Industry, 4=Earnings Date, 5=Earnings Soon
    # 6=Index, 7=Current Price, 8=Date
    # 9=NF Expiry, 10=NF Strike, 11=NF Call Price
    # 12=N2F Expiry, 13=N2F Strike, 14=N2F Call Price
    # 15=NF diff, 16=N2F diff, 17=NF Signal, 18=N2F Signal, 19=Notes

    earnings_warn_fill = PatternFill("solid", start_color="FF0000")  # red bg for earnings soon
    earnings_warn_font_color = "FFFFFF"                              # white text

    for row_idx in range(2, ws.max_row + 1):
        notes     = str(ws.cell(row=row_idx, column=19).value or "")
        index_val = str(ws.cell(row=row_idx, column=6).value or "")
        earn_soon = str(ws.cell(row=row_idx, column=5).value or "")
        has_error = "error" in notes.lower() or "no options" in notes.lower()
        is_both   = index_val == "Both"
        row_fill  = (err_fill if has_error else
                     both_fill if is_both else
                     alt_fill if row_idx % 2 == 0 else PatternFill())

        for col_idx in range(1, len(COLUMNS) + 1):
            cell        = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.font   = Font(name="Arial", size=10)

            if col_idx in (2, 3):                           # Company Name, Industry — left-align
                cell.alignment = left_align
            else:
                cell.alignment = center

            if col_idx == 5:                                # Earnings Soon flag
                if "yes" in earn_soon.lower():
                    cell.fill = earnings_warn_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color=earnings_warn_font_color)
                else:
                    cell.fill = row_fill
            elif col_idx in (11, 14):                       # call price columns
                cell.fill = price_fill
                if cell.value is not None:
                    cell.number_format = "$#,##0.00"
            elif col_idx == 7:                              # current price
                cell.fill = row_fill
                if cell.value is not None:
                    cell.number_format = "$#,##0.00"
            elif col_idx in (15, 16):                       # difference columns
                cell.fill = diff_fill
                if cell.value is not None:
                    cell.number_format = "$#,##0.00;[Red]-$#,##0.00"
            elif col_idx in (17, 18):                       # signal columns
                val = cell.value
                if val == "Good":
                    cell.fill = good_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color="375623")
                elif val == "Bad":
                    cell.fill = bad_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
                else:
                    cell.fill = row_fill
            else:
                cell.fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ["Generated On",      str(date.today())],
        ["Total Tickers",     len(rows)],
        ["With Options Data", sum(1 for r in rows if r["Next Friday Call Price ($)"] is not None)],
        ["Earnings Within 2W",sum(1 for r in rows if r.get("Earnings Soon","") == "⚠️ Yes")],
        ["Errors / No Data",  sum(1 for r in rows if r["Notes"])],
    ]
    for r_idx, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(row=r_idx, column=2, value=value).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    wb.save(str(output_path))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="S&P500 + Dow covered call options puller")
    parser.add_argument("--output",  default="options_output_dt0515.xlsx", help="Output Excel file")
    parser.add_argument("--workers", type=int, default=4, help="Parallel threads (default: 4)")
    args = parser.parse_args()

    output_path = Path(args.output)

    print("=" * 60)
    print("  Stock Options Puller — S&P 500 + Dow Jones")
    print("=" * 60)

    print("\nFetching index constituents...")
    ticker_map = build_ticker_index_map()

    tickers = sorted(ticker_map.keys())
    print(f"\n  Total unique tickers : {len(tickers)}")

    expiry1, expiry2 = get_expiry_dates()
    print(f"\nTarget expiries:")
    print(f"  Next Friday          : {expiry1}")
    print(f"  Next-Next Friday     : {expiry2}")
    print(f"\nProcessing {len(tickers)} tickers with {args.workers} workers...\n")

    rows   = [None] * len(tickers)
    done   = 0
    errors = 0

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        future_to_idx = {
            pool.submit(process_ticker, t, ticker_map[t], expiry1, expiry2): i
            for i, t in enumerate(tickers)
        }
        for future in as_completed(future_to_idx):
            idx       = future_to_idx[future]
            row       = future.result()
            rows[idx] = row
            done     += 1
            if row["Notes"]:
                errors += 1
            pct       = done / len(tickers) * 100
            price_str = f"${row['Current Price']}" if row["Current Price"] else row["Notes"]
            print(f"\r  [{done}/{len(tickers)}] {pct:5.1f}%  {row['Ticker']:<6} {price_str:<20}",
                  end="", flush=True)

    print(f"\n\nComplete. {done - errors} OK  |  {errors} errors/warnings")
    save_excel(rows, output_path)
    print(f"\nDone! Output saved to: {output_path.resolve()}")


if __name__ == "__main__":
    main()