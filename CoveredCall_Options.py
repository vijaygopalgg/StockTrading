"""
stock_options_puller.py
-----------------------
Fetches S&P 500 + Dow Jones tickers, then for each stock pulls:
  - Current price (as of today)
  - ITM Covered Call: 1 strike ABOVE current price
      * Next week Friday expiry
      * Next-to-next week Friday expiry

Ticker source strategy (in order):
  1. Live fetch from Wikipedia (with browser headers)
  2. Fall back to embedded built-in lists if Wikipedia is blocked

Usage:
  python stock_options_puller.py
  python stock_options_puller.py --output "C:\\Users\\vijay\\options_output.xlsx"
  python stock_options_puller.py --workers 8
"""

import argparse
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Embedded ticker lists (fallback) ─────────────────────────────────────────
# Updated as of May 2025. Used only if Wikipedia fetch fails.

DOW_30 = [
    "AAPL", "AMGN", "AMZN", "AXP", "BA", "CAT", "CRM", "CSCO", "CVX", "DIS",
    "GS", "HD", "HON", "IBM", "JNJ", "JPM", "KO", "MCD", "MMM", "MRK",
    "MSFT", "NKE", "NVDA", "PG", "SHW", "TRV", "UNH", "V", "VZ", "WMT",
]

SP500_TICKERS = [
    "A","AAL","AAP","AAPL","ABBV","ABC","ABMD","ABT","ACN","ACGL","ACV","ADI","ADM","ADP","ADSK",
    "AEE","AEP","AES","AFL","AIG","AIZ","AJG","AKAM","ALB","ALGN","ALK","ALL","ALLE","AMAT","AMCR",
    "AMD","AME","AMGN","AMP","AMT","AMZN","ANET","ANSS","AON","AOS","APA","APD","APH","APTV","ARE",
    "ATO","ATVI","AVB","AVGO","AVY","AWK","AXP","AZO","BA","BAC","BAX","BBWI","BBY","BDX","BEN",
    "BF-B","BIIB","BIO","BK","BKNG","BKR","BLK","BMY","BR","BRK-B","BRO","BSX","BWA","BXP","C",
    "CAG","CAH","CARR","CAT","CB","CBOE","CBRE","CCI","CCL","CDNS","CDW","CE","CEG","CF","CFG",
    "CHD","CHRW","CHTR","CI","CINF","CL","CLX","CMA","CMCSA","CME","CMG","CMI","CMS","CNC","CNP",
    "COF","COO","COP","COST","CPB","CPRT","CPT","CRL","CRM","CSCO","CSGP","CSX","CTAS","CTLT",
    "CTRA","CTSH","CTVA","CVS","CVX","CZR","D","DAL","DD","DE","DFS","DG","DGX","DHI","DHR","DIS",
    "DISH","DLR","DLTR","DOV","DOW","DPZ","DRE","DRI","DTE","DUK","DVA","DVN","DXC","DXCM","EA",
    "EBAY","ECL","ED","EFX","EIX","EL","EMN","EMR","ENPH","EOG","EPAM","EQIX","EQR","EQT","ES",
    "ESS","ETN","ETR","EVRG","EW","EXC","EXPD","EXPE","EXR","F","FANG","FAST","FCX","FDS","FDX",
    "FE","FFIV","FIS","FISV","FITB","FLT","FMC","FOX","FOXA","FRC","FRT","FTNT","FTV","GD","GE",
    "GEHC","GEN","GILD","GIS","GL","GLW","GM","GNRC","GOOG","GOOGL","GPC","GPN","GRMN","GS","GWW",
    "HAL","HAS","HBAN","HCA","HD","HES","HIG","HII","HLT","HOLX","HON","HPE","HPQ","HRL","HSIC",
    "HST","HSY","HUM","HWM","IBM","ICE","IDXX","IEX","IFF","ILMN","INCY","INTC","INTU","INVH",
    "IP","IPG","IQV","IR","IRM","ISRG","IT","ITW","IVZ","J","JBHT","JCI","JKHY","JNJ","JNPR",
    "JPM","K","KDP","KEY","KEYS","KHC","KIM","KLAC","KMB","KMI","KMX","KO","KR","L","LDOS","LEN",
    "LH","LHX","LIN","LKQ","LLY","LMT","LNC","LNT","LOW","LRCX","LUV","LVS","LW","LYB","LYV",
    "MA","MAA","MAR","MAS","MCD","MCHP","MCK","MCO","MDLZ","MDT","MET","META","MGM","MHK","MKC",
    "MKTX","MLM","MMC","MMM","MNST","MO","MOH","MOS","MPC","MPWR","MRK","MRNA","MRO","MS","MSCI",
    "MSFT","MSI","MTB","MTD","MU","NCLH","NDAQ","NEE","NEM","NFLX","NI","NKE","NOC","NOV","NOW",
    "NRG","NSC","NTAP","NTRS","NUE","NVDA","NVR","NWL","NWS","NWSA","NXPI","O","OGN","OKE","OMC",
    "ON","ORCL","ORLY","OXY","PAYX","PAYC","PCAR","PCG","PEAK","PEG","PEP","PFE","PFG","PG","PGR",
    "PH","PHM","PKG","PKI","PLD","PM","PNC","PNR","PNW","POOL","PPG","PPL","PRU","PSA","PSX","PTC",
    "PWR","PXD","PYPL","QCOM","QRVO","RCL","RE","REG","REGN","RF","RHI","RJF","RL","RMD","ROK",
    "ROL","ROP","ROST","RSG","RTX","SBAC","SBUX","SCHW","SEE","SHW","SJM","SLB","SNA","SNPS","SO",
    "SPG","SPGI","SRE","STE","STT","STX","STZ","SWK","SWKS","SYF","SYK","SYY","T","TAP","TDG",
    "TDY","TECH","TEL","TER","TFC","TFX","TGT","TJX","TMO","TMUS","TPR","TRMB","TROW","TRV","TSCO",
    "TSLA","TSN","TT","TTWO","TXN","TXT","TYL","UAL","UDR","UHS","ULTA","UNH","UNP","UPS","URI",
    "USB","V","VFC","VLO","VMC","VNO","VRSN","VRTX","VTR","VTRS","VZ","WAB","WAT","WBA","WBD",
    "WDC","WEC","WELL","WFC","WHR","WM","WMB","WMT","WRB","WRK","WST","WTW","WY","WYNN","XEL",
    "XOM","XRAY","XYL","YUM","ZBH","ZBRA","ZION","ZTS",
]


# ── Date helpers ──────────────────────────────────────────────────────────────

def next_friday(from_date=None):
    d = from_date or date.today()
    days_ahead = (4 - d.weekday()) % 7
    if days_ahead == 0:
        days_ahead = 7
    return d + timedelta(days=days_ahead)


def get_expiry_dates():
    f1 = next_friday()
    f2 = next_friday(f1 + timedelta(days=1))
    return str(f1), str(f2)


# ── Ticker fetchers ───────────────────────────────────────────────────────────

def fetch_sp500_live():
    """Try to fetch live S&P 500 list from Wikipedia with browser headers."""
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
        tables = pd.read_html(
            "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
            storage_options={"User-Agent": headers["User-Agent"]},
        )
        tickers = tables[0]["Symbol"].str.upper().str.strip().str.replace(".", "-", regex=False)
        return list(tickers)
    except Exception:
        return None


def fetch_dow_live():
    """Try to fetch live Dow list from Wikipedia with browser headers."""
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        }
        tables = pd.read_html(
            "https://en.wikipedia.org/wiki/Dow_Jones_Industrial_Average",
            storage_options={"User-Agent": headers["User-Agent"]},
        )
        for table in tables:
            for col in table.columns:
                vals = table[col].dropna().astype(str)
                if vals.str.match(r"^[A-Z]{1,5}$").mean() > 0.5:
                    return list(vals.str.upper().str.strip())
        return None
    except Exception:
        return None


def build_ticker_index_map():
    """
    Build {ticker: index_label} map.
    Tries live Wikipedia fetch first; falls back to embedded lists.
    """
    print("  Trying live fetch from Wikipedia...")
    sp500_list = fetch_sp500_live()
    dow_list   = fetch_dow_live()

    if sp500_list:
        print(f"  S&P 500: fetched {len(sp500_list)} tickers (live)")
    else:
        sp500_list = SP500_TICKERS
        print(f"  S&P 500: using embedded list ({len(sp500_list)} tickers)")

    if dow_list:
        print(f"  Dow Jones: fetched {len(dow_list)} tickers (live)")
    else:
        dow_list = DOW_30
        print(f"  Dow Jones: using embedded list ({len(dow_list)} tickers)")

    sp500_set = set(sp500_list)
    dow_set   = set(dow_list)

    combined = {}
    for t in sp500_set:
        combined[t] = "Both" if t in dow_set else "S&P 500"
    for t in dow_set:
        if t not in combined:
            combined[t] = "Dow"

    return combined


# ── Options fetcher ───────────────────────────────────────────────────────────

def get_itm_call(ticker_obj, expiry, current_price):
    """Return (strike, mid_price) for 1 strike above current price."""
    try:
        chain = ticker_obj.option_chain(expiry)
        calls = chain.calls.copy()
        calls = calls[calls["strike"] > current_price].sort_values("strike")
        if calls.empty:
            return None, None
        row    = calls.iloc[0]
        strike = float(row["strike"])
        bid    = float(row["bid"])       if row["bid"]       > 0 else None
        ask    = float(row["ask"])       if row["ask"]       > 0 else None
        last   = float(row["lastPrice"]) if row["lastPrice"] > 0 else None
        if bid and ask:
            mid = round((bid + ask) / 2, 2)
        elif last:
            mid = round(last, 2)
        else:
            mid = None
        return strike, mid
    except Exception:
        return None, None


# ── Per-ticker processing ─────────────────────────────────────────────────────

def process_ticker(ticker, index_label, expiry1, expiry2):
    row = {
        "Ticker": ticker,
        "Index": index_label,
        "Current Price": None,
        "Date": str(date.today()),
        "Next Friday Expiry": expiry1,
        "Next Friday Strike": None,
        "Next Friday Call Price ($)": None,
        "Next-Next Friday Expiry": expiry2,
        "Next-Next Strike": None,
        "Next-Next Call Price ($)": None,
        "NF Strike - Current Price": None,
        "N2F Strike - Current Price": None,
        "NF Signal": None,
        "N2F Signal": None,
        "Notes": "",
    }
    try:
        t     = yf.Ticker(ticker)
        price = round(float(t.fast_info.last_price), 2)
        row["Current Price"] = price

        available = t.options
        if not available:
            row["Notes"] = "No options data"
            return row

        match1 = expiry1 if expiry1 in available else next(
            (e for e in available if e >= expiry1), None)
        if match1:
            row["Next Friday Expiry"] = match1
            row["Next Friday Strike"], row["Next Friday Call Price ($)"] = \
                get_itm_call(t, match1, price)

        match2 = expiry2 if expiry2 in available else next(
            (e for e in available if e > (match1 or expiry1)), None)
        if match2:
            row["Next-Next Friday Expiry"] = match2
            row["Next-Next Strike"], row["Next-Next Call Price ($)"] = \
                get_itm_call(t, match2, price)

        # Derived columns
        nf_strike  = row["Next Friday Strike"]
        n2f_strike = row["Next-Next Strike"]
        if nf_strike is not None:
            diff1 = round(nf_strike - price, 2)
            row["NF Strike - Current Price"] = diff1
            row["NF Signal"] = "Good" if diff1 > 0 else "Bad"
        if n2f_strike is not None:
            diff2 = round(n2f_strike - price, 2)
            row["N2F Strike - Current Price"] = diff2
            row["N2F Signal"] = "Good" if diff2 > 0 else "Bad"

    except Exception as e:
        row["Notes"] = f"Error: {e}"

    return row


# ── Excel output ──────────────────────────────────────────────────────────────

COLUMNS = [
    "Ticker", "Index", "Current Price", "Date",
    "Next Friday Expiry", "Next Friday Strike", "Next Friday Call Price ($)",
    "Next-Next Friday Expiry", "Next-Next Strike", "Next-Next Call Price ($)",
    "NF Strike - Current Price", "N2F Strike - Current Price",
    "NF Signal", "N2F Signal",
    "Notes",
]
COL_WIDTHS = [10, 8, 14, 12, 18, 18, 22, 22, 16, 22, 22, 22, 12, 12, 35]


def save_excel(rows, output_path):
    df = pd.DataFrame(rows)[COLUMNS]
    df.to_excel(str(output_path), index=False)

    wb = load_workbook(str(output_path))
    ws = wb.active

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    alt_fill    = PatternFill("solid", start_color="D6E4F0")
    price_fill  = PatternFill("solid", start_color="E2EFDA")
    both_fill   = PatternFill("solid", start_color="FFF2CC")
    err_fill    = PatternFill("solid", start_color="FCE4D6")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")

    for i in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]
    ws.row_dimensions[1].height = 28

    good_fill = PatternFill("solid", start_color="C6EFCE")   # green  - Good
    bad_fill  = PatternFill("solid", start_color="FFC7CE")   # red    - Bad
    diff_fill = PatternFill("solid", start_color="EDEDED")   # grey   - diff columns

    # Column index map (1-based): Notes is now col 15
    # 11=NF diff, 12=N2F diff, 13=NF Signal, 14=N2F Signal, 15=Notes
    for row_idx in range(2, ws.max_row + 1):
        notes     = str(ws.cell(row=row_idx, column=15).value or "")
        index_val = str(ws.cell(row=row_idx, column=2).value or "")
        has_error = "error" in notes.lower() or "no options" in notes.lower()
        is_both   = index_val == "Both"
        row_fill  = (err_fill if has_error else
                     both_fill if is_both else
                     alt_fill if row_idx % 2 == 0 else PatternFill())

        for col_idx in range(1, len(COLUMNS) + 1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.border    = border
            cell.alignment = center
            cell.font      = Font(name="Arial", size=10)
            if col_idx in (7, 10):                      # call price columns
                cell.fill = price_fill
                if cell.value is not None:
                    cell.number_format = "$#,##0.00"
            elif col_idx == 3:                          # current price
                cell.fill = row_fill
                if cell.value is not None:
                    cell.number_format = "$#,##0.00"
            elif col_idx in (11, 12):                   # difference columns
                cell.fill = diff_fill
                if cell.value is not None:
                    cell.number_format = "$#,##0.00;[Red]-$#,##0.00"
            elif col_idx in (13, 14):                   # signal columns
                val = cell.value
                if val == "Good":
                    cell.fill = good_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color="375623")
                elif val == "Bad":
                    cell.fill = bad_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
                else:
                    cell.fill = row_fill
            else:
                cell.fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ["Generated On",      str(date.today())],
        ["Total Tickers",     len(rows)],
        ["S&P 500 Only",      sum(1 for r in rows if r["Index"] == "S&P 500")],
        ["Dow Only",          sum(1 for r in rows if r["Index"] == "Dow")],
        ["In Both Indexes",   sum(1 for r in rows if r["Index"] == "Both")],
        ["With Options Data", sum(1 for r in rows if r["Next Friday Call Price ($)"] is not None)],
        ["Errors / No Data",  sum(1 for r in rows if r["Notes"])],
    ]
    for r_idx, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(row=r_idx, column=2, value=value).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    wb.save(str(output_path))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="S&P500 + Dow covered call options puller")
    parser.add_argument("--output",  default="options_output.xlsx", help="Output Excel file")
    parser.add_argument("--workers", type=int, default=8, help="Parallel threads (default: 8)")
    args = parser.parse_args()

    output_path = Path(args.output)

    print("=" * 60)
    print("  Stock Options Puller — S&P 500 + Dow Jones")
    print("=" * 60)

    print("\nFetching index constituents...")
    ticker_map = build_ticker_index_map()

    tickers = sorted(ticker_map.keys())
    print(f"\n  Total unique tickers : {len(tickers)}")
    print(f"  S&P 500 only         : {sum(1 for v in ticker_map.values() if v == 'S&P 500')}")
    print(f"  Dow only             : {sum(1 for v in ticker_map.values() if v == 'Dow')}")
    print(f"  In both indexes      : {sum(1 for v in ticker_map.values() if v == 'Both')}")

    expiry1, expiry2 = get_expiry_dates()
    print(f"\nTarget expiries:")
    print(f"  Next Friday          : {expiry1}")
    print(f"  Next-Next Friday     : {expiry2}")
    print(f"\nProcessing {len(tickers)} tickers with {args.workers} workers...\n")

    rows   = [None] * len(tickers)
    done   = 0
    errors = 0

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        future_to_idx = {
            pool.submit(process_ticker, t, ticker_map[t], expiry1, expiry2): i
            for i, t in enumerate(tickers)
        }
        for future in as_completed(future_to_idx):
            idx       = future_to_idx[future]
            row       = future.result()
            rows[idx] = row
            done     += 1
            if row["Notes"]:
                errors += 1
            pct       = done / len(tickers) * 100
            price_str = f"${row['Current Price']}" if row["Current Price"] else row["Notes"]
            print(f"\r  [{done}/{len(tickers)}] {pct:5.1f}%  {row['Ticker']:<6} {price_str:<20}",
                  end="", flush=True)

    print(f"\n\nComplete. {done - errors} OK  |  {errors} errors/warnings")
    save_excel(rows, output_path)
    print(f"\nDone! Output saved to: {output_path.resolve()}")


if __name__ == "__main__":
    main()